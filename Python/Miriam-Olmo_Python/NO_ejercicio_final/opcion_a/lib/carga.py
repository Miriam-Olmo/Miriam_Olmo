import csv
import json
import xml.etree.ElementTree as et
from openpyxl import load_workbook

def cargar_csv(ruta):
    fichero = open(ruta, "r", encoding='UTF-8')
    lector = csv.DictReader(fichero) 
    artistas = list(lector)
    fichero.close()
    return artistas

def cargar_excel(ruta):
    excel = load_workbook(ruta, data_only=True)
    hoja = excel.active
    
    filas = hoja.iter_rows(values_only=True)
    cabeceras = next(filas)
    
    lista_resultante = []
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if not any(fila):# Evita almacenar líneas huérfanas o vacías
            continue
        resultado = dict(zip(cabeceras, fila))
        lista_resultante.append(resultado)
       
    return lista_resultante

def cargar_json(ruta):
    fichero = open(ruta, "r", encoding="UTF-8")
    datos = json.load(fichero)
    fichero.close()
    return datos

def cargar_xml(ruta):
    fichero = et.parse(ruta)
    nodo_raiz = fichero.getroot()

    patrocinadores = []
    for patrocinador in nodo_raiz.findall('patrocinador'):
        datos_patrocinador = {
            'nombre_empresa': patrocinador.find('nombre_empresa').text,
            'contacto': patrocinador.find('contacto').text,
            'email': patrocinador.find('email').text,
            'importe_patrocinio': patrocinador.find('importe_patrocinio').text,
            'categoria': patrocinador.find('categoria').text,
            'fecha_inicio': patrocinador.find('fecha_inicio').text,
            'fecha_fin': patrocinador.find('fecha_fin').text
        }
        patrocinadores.append(datos_patrocinador)
        
    return patrocinadores


