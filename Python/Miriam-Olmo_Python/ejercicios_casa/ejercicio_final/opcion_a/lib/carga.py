# importar csv
import csv
#importar json
import json
# importar xml
import xml.etree.ElementTree as et
# importar excel para su lectura
from openpyxl import load_workbook

# cargar csv 
def cargar_csv(ruta):
    fichero = open(ruta, 'r', encoding='utf-8')
    lector = csv.DictReader(fichero)
    artistas = list(lector)
    fichero.close()
    return artistas

def cargar_json(ruta):
    fichero = open(ruta, 'r', encoding='utf-8')
    datos = json.load(fichero)
    fichero.close()
    return datos

def cargar_xml(ruta):
    arbol = et.parse(ruta)
    nodo_padre = arbol.getroot()
    patrocinadores = []

    for patrocinador in nodo_padre.findall('patrocinador'):
        datos_patrocinador ={
            'nombre_empresa': patrocinador.find('nombre_empresa').text,
            'contacto': patrocinador.find('contacto').text,
            'email': patrocinador.find('email').text,
            'importe_patrocinio': patrocinador.find('importe_patrocinio').text if patrocinador.find is not None else None,
            'categoria': patrocinador.find('categoria').text,
            'fecha_inicio': patrocinador.find('fecha_inicio').text,
            'fecha_fin': patrocinador.find('fecha_fin').text
        }
        patrocinadores.append (datos_patrocinador)
    return patrocinadores

def cargar_excel(ruta):
    excel = load_workbook(ruta)
    hoja = excel.active
    filas = hoja.iter_rows(values_only=True)
    cabeceras = next(filas)
    lista = []
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        dato = dict(zip(cabeceras, fila))
        lista.append(dato)
    return lista