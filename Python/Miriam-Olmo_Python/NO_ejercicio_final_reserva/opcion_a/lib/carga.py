# ============================================
# LIB/CARGA.PY
# ============================================

import csv
import json

import xml.etree.ElementTree as ElementTree

from openpyxl import load_workbook

# ============================================
# CARGAR CSV
# ============================================

def cargar_csv(ruta_archivo):

    lista_registros = []

    archivo = open(
        ruta_archivo,
        mode="r",
        encoding="utf-8"
    )

    try:

        lector_csv = csv.DictReader(
            archivo
        )

        for fila in lector_csv:

            lista_registros.append(
                dict(fila)
            )

    finally:

        archivo.close()

    return lista_registros

# ============================================
# CARGAR EXCEL
# ============================================

def cargar_excel(
    ruta_archivo,
    nombre_hoja=None
):

    lista_registros = []

    libro_excel = load_workbook(
        ruta_archivo
    )

    if nombre_hoja is not None:

        hoja_excel = libro_excel[
            nombre_hoja
        ]

    else:

        hoja_excel = libro_excel.active

    encabezados = [

        celda.value

        for celda in hoja_excel[1]
    ]

    for fila in hoja_excel.iter_rows(

        min_row=2,

        values_only=True
    ):

        diccionario_fila = dict(

            zip(
                encabezados,
                fila
            )
        )

        lista_registros.append(
            diccionario_fila
        )

    return lista_registros

# ============================================
# CARGAR JSON
# ============================================

def cargar_json(ruta_archivo):

    archivo = open(
        ruta_archivo,
        mode="r",
        encoding="utf-8"
    )

    try:

        lista_registros = json.load(
            archivo
        )

    finally:

        archivo.close()

    return lista_registros

# ============================================
# CARGAR XML
# ============================================

def cargar_xml(ruta_archivo):

    lista_registros = []

    archivo = open(
        ruta_archivo,
        mode="r",
        encoding="utf-8"
    )

    try:

        arbol_xml = ElementTree.parse(
            archivo
        )

        raiz_xml = arbol_xml.getroot()

        for elemento in raiz_xml:

            registro = {}

            for campo in elemento:

                registro[
                    campo.tag
                ] = campo.text

            lista_registros.append(
                registro
            )

    finally:

        archivo.close()

    return lista_registros